"""
Генератор протоколов испытания в формате .xlsx.

Поддерживаемые типы:
  - gost_58401  (АБС / ЩМАС, ГОСТ Р 58401)

Добавление нового типа: добавить ветку в generate() и новый модуль layout.
"""
from __future__ import annotations

from io import BytesIO
from typing import TYPE_CHECKING

from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_to_tuple, column_index_from_string

from app import config as app_config
from app.services import protocol_abs58401_layout as L

if TYPE_CHECKING:
    from app.models.client import Client
    from app.models.protocol import Protocol
    from app.models.request import Request
    from app.models.sample import Sample
    from app.models.test import Test


def generate(
    protocol: "Protocol",
    test: "Test",
    sample: "Sample",
    client: "Client | None",
    request: "Request | None",
) -> BytesIO:
    """Выбирает генератор по типу испытания и возвращает заполненный BytesIO."""
    t = (test.test_type or "").lower()
    if "58401" in t or "58406" in t:
        return _generate_abs58401(protocol, test, sample, client, request)
    raise ValueError(f"Генератор для типа испытания '{test.test_type}' не реализован")


def _unmerged_cell(ws, row: int, col: int):
    """Возвращает верхнюю левую ячейку объединения (или саму ячейку если не объединена)."""
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return ws.cell(rng.min_row, rng.min_col)
    return ws.cell(row, col)


def _write(ws, cell_ref: str, value):
    """Запись в ячейку с учётом объединений."""
    if value is None:
        return
    row, col = coordinate_to_tuple(cell_ref)
    _unmerged_cell(ws, row, col).value = value


def _write_rc(ws, row: int, col: int, value):
    """Запись по номеру строки/столбца с учётом объединений."""
    if value is None:
        return
    _unmerged_cell(ws, row, col).value = value


# ---------------------------------------------------------------------------
# ГОСТ Р 58401 / 58406 — АБС, ЩМАС
# ---------------------------------------------------------------------------

def _generate_abs58401(
    protocol: "Protocol",
    test: "Test",
    sample: "Sample",
    client: "Client | None",
    request: "Request | None",
) -> BytesIO:
    tpl_path = app_config.protocol_template_path("gost_58401")
    if tpl_path is None or not tpl_path.is_file():
        raise FileNotFoundError(
            f"Шаблон протокола не найден: {tpl_path}. "
            "Загрузите файл через раздел «Протоколы» → «Загрузить шаблон»."
        )

    wb = load_workbook(str(tpl_path))

    if L.SHEET_ABS not in wb.sheetnames:
        raise ValueError(
            f"В шаблоне нет листа «{L.SHEET_ABS}». Проверьте файл шаблона."
        )
    ws = wb[L.SHEET_ABS]

    # --- Шапка ---
    _write(ws, L.CELL_PROTOCOL_NUMBER, f"{protocol.number:05d}")
    if protocol.created_at:
        _write(ws, L.CELL_PROTOCOL_DATE, protocol.created_at.strftime("%d.%m.%Y"))

    # --- Общие сведения ---
    if client:
        _write_rc(ws, L.ROW_CUSTOMER, L.GENERAL_VALUE_COL, client.name)
        contact = " / ".join(filter(None, [client.contact_phone, client.contact_email]))
        _write_rc(ws, L.ROW_CONTACT, L.GENERAL_VALUE_COL, contact or None)

    _write_rc(ws, L.ROW_MANUFACTURER,         L.GENERAL_VALUE_COL, sample.manufacturer)
    _write_rc(ws, L.ROW_SAMPLE_SOURCE,         L.GENERAL_VALUE_COL, sample.sampled_by)
    obj_name = f"{sample.material_name or ''} {sample.material_grade or ''}".strip()
    _write_rc(ws, L.ROW_TEST_OBJECT,           L.GENERAL_VALUE_COL, obj_name or None)
    if request:
        _write_rc(ws, L.ROW_CONSTRUCTION_OBJECT, L.GENERAL_VALUE_COL, request.notes or request.number)
    _write_rc(ws, L.ROW_SAMPLING_PLACE,        L.GENERAL_VALUE_COL, sample.sampling_location)
    if sample.registration_date:
        _write_rc(ws, L.ROW_REGISTRATION_DATE, L.GENERAL_VALUE_COL,
                  sample.registration_date.strftime("%d.%m.%Y"))
    if test.tested_at:
        _write_rc(ws, L.ROW_TEST_DATES, L.GENERAL_VALUE_COL, test.tested_at.strftime("%d.%m.%Y"))
    _write_rc(ws, L.ROW_MIX_TYPE, L.GENERAL_VALUE_COL, sample.material_grade)

    # --- Физ-мех показатели ---
    def _result(row: int, value):
        if value is None:
            return
        col = column_index_from_string(L.COL_RESULT)
        try:
            _unmerged_cell(ws, row, col).value = float(value)
        except (TypeError, ValueError):
            _unmerged_cell(ws, row, col).value = value
    bulk   = (test.bulk_density   or {}).get("avg")
    maxd   = (test.max_density    or {}).get("avg")
    voids  = test.air_voids
    binder = (test.binder_content or {}).get("avg")

    _result(L.ROW_BULK_DENSITY,   bulk)
    _result(L.ROW_MAX_DENSITY,    maxd)
    _result(L.ROW_AIR_VOIDS,      voids)
    _result(L.ROW_BINDER_CONTENT, binder)

    # --- Зерновой состав ---
    grain = test.grain_composition or {}
    avg_vals = grain.get("avg", [])
    for col_offset, val in enumerate(avg_vals):
        col_idx = L.FIRST_SIEVE_COL + col_offset
        if col_idx > L.LAST_SIEVE_COL:
            break
        if val not in (None, "", "-"):
            try:
                _unmerged_cell(ws, L.ROW_GRAIN_ACTUAL, col_idx).value = float(val)
            except (TypeError, ValueError):
                _unmerged_cell(ws, L.ROW_GRAIN_ACTUAL, col_idx).value = val

    # --- Заключение ---
    if protocol.conclusion:
        _write(ws, L.CELL_CONCLUSION_TEXT, protocol.conclusion)

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out
